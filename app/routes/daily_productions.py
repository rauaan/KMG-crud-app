"""Маршруты для управления суточными производственными рапортами.

Модуль содержит обработчики запросов для просмотра, создания,
редактирования и удаления производственных рапортов скважин.
"""

import io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from app.extensions import db
from app.models import DailyProduction, Well
from app.forms import CreateDailyProduction
from openpyxl import Workbook, load_workbook

daily_productions_bp = Blueprint("daily_productions", __name__)


@daily_productions_bp.route("/")
@login_required
def daily_productions():
    """Отображает список производственных рапортов.

    Получает все рапорты из базы данных и передает их
    в HTML-шаблон для отображения.

    Returns:
        Response: Страница со списком рапортов.
    """

    if request.method == "GET":
        reports = DailyProduction.query.all()
    return render_template("main/list_daily_productions.html", reports=reports)


@daily_productions_bp.route("/create", methods=["GET", "POST"])
@login_required
def create_daily_production():
    """Создает новый производственный рапорт.

    При GET-запросе отображает форму создания рапорта.
    При POST-запросе выполняет валидацию данных и сохраняет
    новый рапорт в базе данных.

    Returns:
        Response: Форма создания или перенаправление
        к списку рапортов.
    """

    form = CreateDailyProduction()
    if form.validate_on_submit():
        new_daily_production = DailyProduction(
            well_id=form.well_id.data,
            date=form.date.data,
            operating_hours=form.operating_hours.data,
            liquid_produced=form.liquid_produced.data,
            water_cut=form.water_cut.data,
            density=form.density.data,
        )
        try:
            db.session.add(new_daily_production)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"ERROR{e}"

        return redirect(url_for("daily_productions.daily_productions"))

    for field, errors in form.errors.items():
        for error in errors:
            flash(error, "warning")

    return render_template(
        "main/create_daily_production.html",
        form=form,
        title="Создать рапорт",
        button_text="Создать",
    )


@daily_productions_bp.route("/edit/<int:well_id>/<date>", methods=["GET", "POST"])
@login_required
def edit_daily_production(well_id, date):
    """Редактирует существующий производственный рапорт.

    Рапорт идентифицируется составным первичным ключом,
    состоящим из идентификатора скважины и даты.

    Args:
        well_id: Идентификатор скважины.
        date: Дата производственного рапорта.

    Returns:
        Response: Форма редактирования или перенаправление
        к списку рапортов после успешного сохранения.
    """

    report = DailyProduction.query.filter_by(well_id=well_id, date=date).first_or_404()

    form = CreateDailyProduction(
        obj=report, original_well_id=report.well_id, original_date=report.date
    )

    if form.validate_on_submit():
        report.well_id = form.well_id.data
        report.date = form.date.data
        report.operating_hours = form.operating_hours.data
        report.liquid_produced = form.liquid_produced.data
        report.water_cut = form.water_cut.data
        report.density = form.density.data

        try:
            db.session.commit()
            return redirect(url_for("daily_productions.daily_productions"))

        except Exception as e:
            db.session.rollback()
            return f"ERROR {e}"

    else:
        for errors in form.errors.values():
            for error in errors:
                flash(error, "warning")

    return render_template(
        "main/create_daily_production.html",
        form=form,
        title="Редактировать рапорт",
        button_text="Сохранить",
    )


@daily_productions_bp.route("/delete/<int:well_id>/<date>", methods=["POST"])
@login_required
def delete_daily_production(well_id, date):
    """Удаляет производственный рапорт.

    Args:
        well_id: Идентификатор скважины.
        date: Дата производственного рапорта.

    Returns:
        Response: Перенаправление к списку рапортов.
    """

    to_delete = DailyProduction.query.filter_by(
        well_id=well_id, date=date
    ).first_or_404()

    try:
        db.session.delete(to_delete)
        db.session.commit()
        return redirect(url_for("daily_productions.daily_productions"))

    except Exception as e:
        return f"ERROR {e}"

    
@daily_productions_bp.route("/download", methods=["GET"])
@login_required
def download_daily_productions():
    """Экспортирует рапорты из базы данных в Excel-файл.

    Позволяет указать начальную и конечную дату для выгрузки
    суточных производственных рапортов. Если даты не указаны,
    выгружаются все рапорты.

    Returns:
        Response: XLSX-файл с рапортами за выбранный период.
    """

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    query = DailyProduction.query

    if date_from:
        date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        query = query.filter(DailyProduction.date >= date_from)

    if date_to:
        date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        query = query.filter(DailyProduction.date <= date_to)

    reports = (
        query
        .order_by(DailyProduction.date)
        .all()
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Рапорты"

    headers = [
        "Скважина (ID)",
        "Дата",
        "Время работы",
        "Жидкость",
        "Обводненность",
        "Плотность",
        "Чистая нефть",
    ]

    sheet.append(headers)

    for report in reports:
        sheet.append([
            report.well_id,
            report.date,
            report.operating_hours,
            report.liquid_produced,
            report.water_cut,
            report.density,
            report.net_oil,
        ])

    # Форматирование заголовков
    for cell in sheet[1]:
        cell.font = cell.font.copy(bold=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    # Ширина колонок
    widths = [15, 15, 18, 15, 18, 15, 18]

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    # Создаём файл в памяти
    output = io.BytesIO()

    workbook.save(output)
    output.seek(0)

    if date_from and date_to:
        filename = (
            f"daily_productions_"
            f"{date_from}_{date_to}.xlsx"
        )
    elif date_from:
        filename = f"daily_productions_from_{date_from}.xlsx"
    elif date_to:
        filename = f"daily_productions_to_{date_to}.xlsx"
    else:
        filename = "daily_productions_all.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )

@daily_productions_bp.route("/upload", methods=["GET", "POST"])
@login_required
def upload_daily_productions():
    """Импортирует суточные рапорты из Excel-файла.

    Ожидает файл в том же формате, в котором данные выгружаются
    через download_daily_productions: первая строка - заголовки,
    далее строки с колонками well_id, date, operating_hours,
    liquid_produced, water_cut, density (колонка "Чистая нефть"
    игнорируется, так как вычисляется автоматически).

    Строки с ошибками (несуществующая скважина, дублирующийся
    рапорт, некорректные значения) пропускаются, об этом сообщается
    через flash-сообщения. Валидные строки сохраняются одной
    транзакцией.

    Returns:
        Response: Перенаправление к списку рапортов с
        flash-сообщением об итогах импорта.
    """

    if request.method == "GET":
        return redirect(url_for("daily_productions.daily_productions"))

    uploaded_file = request.files.get("file")

    if uploaded_file is None or uploaded_file.filename == "":
        flash("Файл не выбран", "warning")
        return redirect(url_for("daily_productions.daily_productions"))

    if not uploaded_file.filename.lower().endswith(".xlsx"):
        flash("Допустим только формат .xlsx", "warning")
        return redirect(url_for("daily_productions.daily_productions"))

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
    except Exception:
        flash("Не удалось прочитать файл. Проверьте формат.", "warning")
        return redirect(url_for("daily_productions.daily_productions"))

    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # пропускаем заголовок

    if not rows:
        flash("Файл не содержит данных", "warning")
        return redirect(url_for("daily_productions.daily_productions"))

    existing_wells = {w.id for w in Well.query.all()}
    existing_reports = {
        (dp.well_id, dp.date)
        for dp in DailyProduction.query.with_entities(
            DailyProduction.well_id, DailyProduction.date
        ).all()
    }

    errors = []
    to_create = []
    seen_in_file = set()

    for index, row in enumerate(rows, start=2):
        if row is None or all(cell is None for cell in row):
            continue  # пустая строка

        if len(row) < 6:
            errors.append(f"Строка {index}: недостаточно столбцов")
            continue

        well_id_raw, date_raw, hours_raw, liquid_raw, water_cut_raw, density_raw = row[:6]

        if well_id_raw is None:
            errors.append(f"Строка {index}: не указан ID скважины")
            continue

        try:
            well_id = int(well_id_raw)
        except (TypeError, ValueError):
            errors.append(f"Строка {index}: некорректный ID скважины")
            continue

        if isinstance(date_raw, datetime):
            date_value = date_raw.date()
        elif hasattr(date_raw, "year") and hasattr(date_raw, "month"):
            # уже date-подобный объект (openpyxl может вернуть datetime.date напрямую)
            date_value = date_raw
        elif isinstance(date_raw, str):
            try:
                date_value = datetime.strptime(date_raw, "%Y-%m-%d").date()
            except ValueError:
                errors.append(f"Строка {index}: некорректный формат даты")
                continue
        else:
            errors.append(f"Строка {index}: не указана дата")
            continue

        try:
            operating_hours = float(hours_raw)
            liquid_produced = float(liquid_raw)
            water_cut = float(water_cut_raw)
            density = float(density_raw)
        except (TypeError, ValueError):
            errors.append(f"Строка {index}: числовые поля заполнены некорректно")
            continue

        if not (0 <= operating_hours <= 24):
            errors.append(f"Строка {index}: время работы должно быть от 0ч до 24ч")
            continue

        if not (0 <= water_cut <= 100):
            errors.append(f"Строка {index}: обводненность должна быть от 0%% до 100%%")
            continue

        if well_id not in existing_wells:
            errors.append(f"Строка {index}: скважина с ID {well_id} не существует")
            continue

        key = (well_id, date_value)

        if key in existing_reports:
            errors.append(
                f"Строка {index}: рапорт для скважины {well_id} за {date_value} уже существует"
            )
            continue

        if key in seen_in_file:
            errors.append(
                f"Строка {index}: дублирующаяся запись внутри файла "
                f"(скважина {well_id}, дата {date_value})"
            )
            continue

        seen_in_file.add(key)
        to_create.append(
            DailyProduction(
                well_id=well_id,
                date=date_value,
                operating_hours=operating_hours,
                liquid_produced=liquid_produced,
                water_cut=water_cut,
                density=density,
            )
        )

    if to_create:
        try:
            db.session.add_all(to_create)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return f"ERROR {e}"

        flash(f"Импортировано рапортов: {len(to_create)}", "success")

    if errors:
        error_summary = "; ".join(errors[:10])
        if len(errors) > 10:
            error_summary += f" и еще {len(errors) - 10} ошибок"
        flash(f"Пропущено строк: {len(errors)}. {error_summary}", "warning")

    return redirect(url_for("daily_productions.daily_productions"))